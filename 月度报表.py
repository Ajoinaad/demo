#!/usr/bin/env python
# coding: utf-8
import pandas as pd
import numpy as np
import os
import warnings
import openpyxl
import openpyxl.styles
import datetime
import openpyxl
import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
import csv
import shutil

from openpyxl import Workbook,load_workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
import pandas as pd
warnings.filterwarnings('ignore')




osList = os.listdir('./input/')
if not osList:
    print('没有这个文件')
    exit()

def open_file():
    files = r"input"  # 打开文件夹
    lists = os.listdir(files)  # 列出目录的下所有文件和文件夹保存到lists
    lists.sort(key=lambda fn: os.path.getmtime(files + "\\" + fn))  # 按时间排序
    file_new = os.path.join(files, lists[-1])
    return file_new


def dealwith():
    print('\n\n\n')
    print('数据正在处理中...')
    # 0.删除旧文件,创建新文件
    name = ' '
    try:
        shutil.rmtree(rf'D:\工作源文件')
    except:
        print('数据处理出现异常,请关闭所有文件夹/文件,并检查D:\盘是否有D:\工作源文件,有则删除,没有请重新运行')

    os.makedirs(rf'D:\工作源文件')
    # 1.读取最新需要处理文件
    df_deal = pd.read_excel(f'./{open_file()}')

    # 1.1重复值处理(目前还不确定是否要弄)
    # df_deal = df.drop_duplicates(subset=['有效证件号','是否排除密接/次密','密接类型','转归'], keep='first', inplace=True)
    # 2.数据空值处理
    df_deal['审核时间'].fillna('空白', inplace=True)
    df_deal['镇（街道）'].fillna('空白', inplace=True)
    df_deal['转出目的省(直辖市)'].fillna(df_deal['目前所处位置'], inplace=True)
    df_deal['转出目的省(直辖市)'].fillna(df_deal['现住址'], inplace=True)
    df_deal['转出目的省(直辖市)'].fillna('不明', inplace=True)
    df_deal['转归'].fillna('空白', inplace=True)
    df_deal['医学观察场所名称'].fillna('空白', inplace=True)
    # 街道错误处理
    road_List = ['同德', '松洲', '黄石', '石井', '鹤龙', '龙归', '金沙', '嘉禾', '云城', '三元里', '京溪', '同和', '人和', '均禾', '大源', '太和',
                 '白云湖', '景泰', '棠景', '永平', '江高', '石门', '新市', '钟落潭']
    name_List = name.split()
    # name_List = name.split('、')
    name_type = df_deal['关联病例'].str.contains('|'.join(name_List))
    must = (df_deal['是否排除密接/次密'] != '是') & (df_deal['是否追踪到'] != '转出外省') & (df_deal['地市'] == '广州市') & (
            df_deal['区县'] == '白云区') & name_type
    worng_must = must & (df_deal['镇（街道）'].str.contains('|'.join(road_List)) == False)
    df_deal['镇（街道）'].mask(worng_must, '待甄别街道', inplace=True)
    # 2.1转换时间控制处理:待转运处理
    df_deal['一码通入住日期待转运'] = df_deal['一码通入住日期'].fillna('空白')
    df_deal['医学观察方式'].mask(((df_deal['一码通入住日期待转运'] == '空白') & (df_deal['居家隔离原因'].isnull())), '待转运',
                           inplace=True)
    # 2.1待转运处理（1018更新）
    df_deal['医学观察方式'].mask(df_deal['医学观察场所名称'].str.contains('待转运'), '待转运', inplace=True)

    # 2.2日期处理
    df_deal['一码通入住日期'].fillna('1970-12-30 23:59', inplace=True)
    df_deal['最后接触日期'].fillna('1970-12-31 23:59', inplace=True)
    df_deal['一码通入住日期'] = pd.to_datetime(df_deal['一码通入住日期']).dt.floor('d')
    df_deal['最后接触日期'] = pd.to_datetime(df_deal['最后接触日期']).dt.floor('d')
    # 2.3数据空值预处理
    df_deal['审核时间'].fillna('空白', inplace=True)
    df_deal['镇（街道）'].fillna('空白', inplace=True)
    df_deal['是否核心密接'].fillna('否', inplace=True)
    # 2.4集中处理
    # 最后接触日期<一码通入住日期就改为集中
    df_deal['医学观察方式'].mask(((df_deal['最后接触日期']) < (df_deal['一码通入住日期'])), '集中', inplace=True)
    df_deal['医学观察方式'].mask(((df_deal['医学观察场所名称'].str.contains('酒店')) & (df_deal['目前所处位置'] != df_deal['医学观察场所名称'])),
                               '集中', inplace=True)
    # 375个场所

    # 3.确诊处理
    df_deal['转归'] = df_deal['转归'].str.replace(pat='.*确诊.*', repl='转为确诊', regex=True)
    df_deal['转归'] = df_deal['转归'].str.replace(pat='.*阳性.*', repl='转为确诊', regex=True)

    # 4.10月5日更新：次密：医学观察方式为“待转运”皆归为“居家”（并把居家隔离原因里的待转运删除）
    df_deal['医学观察方式'].mask((df_deal['医学观察方式'].str.contains('待转运')) & (df_deal['密接类型'] == '密接的密接'), '居家',
                           inplace=True)
    df_deal['居家隔离原因'].mask((df_deal['医学观察方式'].str.contains('待转运')) & (df_deal['密接类型'] == '密接的密接'), '空白',
                               inplace=True)

    df_deal = df_deal.get(
        ['地市', '区县', '镇（街道）', 'ID', '姓名', '国籍', '性别', '年龄', '有效证件号', '联系方式', '目前所处位置', '现住址', '职业', '工作单位',
         '密接/次密发现途径', '是否核心密接', '是否排除密接/次密', '关联病例', '关联密接', '密接类型', '与患者关系', '接触地点', '最后接触日期', '应解除观察日期', '关联重点场所',
         '转归', '备注', '是否追踪到', '审核时间', '录入时间', '医学观察方式', '创建单位', '转出目的省(直辖市)', '居家隔离原因'])

    # 加快速度专用！这样排除密接/次密身份!!!!!(这一句大疫情过后可以删除)
    df_deal = df_deal[df_deal['是否排除密接/次密'] != '是']
    # 新模板删除次密部分（11.12日以后）
    df_deal = df_deal[df_deal['密接类型'] != '密接的密接']

    #
    df_deal.to_csv(f'D:/工作源文件/{Original().open_file()[5:-5]}.csv', index=False)  # 保存最新版数据
    # df_deal.to_excel(f'D:/工作源文件/{Original().open_file()[5:]}', index=False)  # 保存最新版数据
    print('数据处理完成!')

def report_newusual():
    data_chart2 = {}
    # 读取处理文件
    # df = pd.read_csv(f'D:/工作源文件/{Original().open_file()[5:-5]}.csv')
    df = pd.read_excel(f'{open_file()}')
    name = ' '
    # 名字类别
    name_List = name.split()
    name_type = df['关联病例'].str.contains('|'.join(name_List))


    #数据处理

    df['审核时间'].fillna('空白', inplace=True)
    # 是否核心密接
    # 先计算密接类型【核密or一般密接】

    type = ['是','']
    sick_type = '核心密接'

    for i in range(0, len(type)):
        # 插入:统计第四点:先统计第四点
        # 第一点
        # 累计甄别
        count = (df['是否排除密接/次密'] != '是') & (df['是否核心密接'].str.contains('|'.join(type[i]))) & name_type

        # 我区主动甄别
        active = count & (df['创建单位'] == '白云区疾病预防控制中心')
        active_count = len(df[active])

        # 推送外省管控人数
        province = active & (df['是否追踪到'] == '转出外省')
        province_count = len(df[province])

        # 推送外市管控人数
        city = active & (df['是否追踪到'] != '转出外省') & (df['地市'] != '广州市')
        city_count = len(df[city])

        # 推送外区管控人数
        area = active & (df['是否追踪到'] != '转出外省') & (df['地市'] == '广州市') & (df['区县'] != '白云区')
        area_count = len(df[area])

        local = active & (df['是否追踪到'] != '转出外省') & (df['地市'] == '广州市') & (df['区县'] == '白云区')
        local_count = len(df[local])

        # 区外推送直接计算
        pushme_count = len(df[count]) - len(df[active])

        # 外区推送我区,我区推送外省(求Y的值) :区外推送X人（推送外省'Y'人）
        out = count & (df['创建单位'] != '白云区疾病预防控制中心') & (df["是否追踪到"] == "转出外省")
        out_count = len(df[out])

        # 第二点
        # 涉及我区应管
        must = count & (df['是否追踪到'] != '转出外省') & (df['地市'] == '广州市') & (df['区县'] == '白云区')
        must_count = len(df[must])

        # 已落地
        workable = (must & (df['审核时间'] != '空白')) + (must & (df['转归'] == '转为确诊'))
        workable_count = len(df[workable])

        # 目前在管
        atpresent = workable & (df['转归'].str.contains('继续'))
        atpresent_count = len(df[atpresent])


        # 集中
        focus = atpresent & (df['医学观察方式'].str.contains('集中'))
        focus_count = len(df[focus])
        data_chart2[f'{sick_type}集中'] = focus_count

        # 待转运
        wait = atpresent & (df['医学观察方式'].str.contains('待转运'))
        wait_count = len(df[wait])
        data_chart2[f'{sick_type}待转运'] = wait_count

        # 居家
        home = atpresent & (df['医学观察方式'].str.contains('居家'))
        home_count = len(df[home])
        data_chart2[f'{sick_type}居家'] = home_count

        # 医院隔离
        hospital = atpresent & (df['医学观察方式'].str.contains('医院'))
        hospital_count = len(df[hospital])
        data_chart2[f'{sick_type}医院隔离'] = hospital_count

        # 解除隔离
        remove = must & (df['转归'].str.contains('解除')) & (df['审核时间'] != '空白')
        remove_count = len(df[remove])
        data_chart2[f'{sick_type}解除观察'] = remove_count

        # 正在核实追踪
        track = must & (df['审核时间'] == '空白') & (df['转归'] != '转为确诊')
        track_count = len(df[track])
        data_chart2[f'{sick_type}核实追踪'] = track_count
        # print(track_count)

        # 转为确诊
        sun = must & (df['转归'] == '转为确诊')
        sun_count = len(df[sun])
        data_chart2[f'{sick_type}转为确诊'] = sun_count

        # 第三点，累计已推送区外管控人数（直接计算）
        allpush_count = province_count + city_count + area_count + out_count
        allpush_str = f'3.累计已推送区外管控{allpush_count}人，已完成双握手；'


        sick_type = '密接'
    # print(data_chart2)
    return data_chart2



#专项名称
excel_name = f'截止{open_file()[6:-5]}，库中数据追踪表'

def statistics():
    df = pd.read_excel(f'{open_file()}')
    # df = pd.read_csv(f'D:\统计表项目(8月25日)\统计表项目(8月25日)\汇报总项\可视化\源文件{open_file()[5:-5]}.csv')
    #报表样式
    data_col = pd.DataFrame(columns=['开始时间','病例（专项名称）','累计甄别核心密接','主动甄别核心密接','推送外区核心密接','推送外市核心密接','推送外省核心密接','区外推送核心密接','我区应管核心密接','核心密接已落地','核心密接在管','核心密接集中','核心密接待转运','核心密接居家','核心密接医院隔离','核心密接解除隔离','核心密接转为确诊','核心密接核实追踪','核心密接新增人数',
    '累计甄别密接','主动甄别密接','推送外区密接','推送外市密接','推送外省密接','区外推送密接','我区应管密接','密接已落地','密接在管','密接集中','密接待转运','密接居家','密接医院隔离','密接解除隔离','密接转为确诊','密接核实追踪','密接新增人数'])
    data_col.set_index('开始时间',inplace=True)
    if not os.path.exists(rf'./output/{excel_name}.xlsx'):
        data_col.to_excel(rf'./output/{excel_name}.xlsx',sheet_name='数据')
        wb = openpyxl.load_workbook(rf'./output/{excel_name}.xlsx')
        # 按索引打开sheet表
        sheet = wb['数据']
        # 填充样式
        file = openpyxl.styles.PatternFill("solid", fgColor="00b0f0")
        file2 = openpyxl.styles.PatternFill("solid", fgColor="1a7fc6")
        file3 = openpyxl.styles.PatternFill("solid", fgColor="00b050")
        # 对指定行列进行颜色填充
        for i in range(1, 3):
            sheet.cell(row=1, column=i).fill = file
        for i in range(3, 20):
            sheet.cell(row=1, column=i).fill = file2
        for i in range(20, 37):
            sheet.cell(row=1, column=i).fill = file3
        wb.save(rf'./output/{excel_name}.xlsx')
    else:
        pass

    data_list = []

    # 病例
    name = ' '
    name_list = name.split(' ')
    name_type = df['关联病例'].str.contains('|'.join(name_list))
    #########密接#######
    #累计甄别密接
    mijie_leiji = (df['是否排除密接/次密'] != '是') & (df['密接类型'] == '密切接触者') & name_type

    #我区主动甄别密接
    mijie_active = mijie_leiji & (df['创建单位'] == '白云区疾病预防控制中心')

    #推送外省
    mijie_pro = mijie_active & (df['是否追踪到'] == '转出外省')

    #推送外市
    mijie_outcity = mijie_active & (df['是否追踪到'] != '转出外省') & (df['地市'] != '广州市')

    #推送外区
    mijie_waiqu = mijie_active & (df['是否追踪到'] != '转出外省') & (df['地市'] == '广州市') & (df['区县'] != '白云区')

    #区外推送
    mijie_quwai = mijie_leiji & (df['创建单位'] != '白云区疾病预防控制中心')

    #我区应管
    mijie_yingguan = mijie_leiji & (df['地市'] == '广州市') & (df['区县'] == '白云区') & (df['是否追踪到'] != '转出外省')



    #########核心密接#######
    #累计甄别核心密接
    hexinmijie_leiji = (df['是否排除密接/次密'] != '是') & (df['密接类型'] == '密切接触者') & name_type & (df['是否核心密接'] == '是')

    #我区主动甄别核心密接
    hexinmijie_active = hexinmijie_leiji & (df['创建单位'] == '白云区疾病预防控制中心')

    #推送外省
    hexinmijie_pro = hexinmijie_active & (df['是否追踪到'] == '转出外省')

    #推送外市
    hexinmijie_outcity = hexinmijie_active & (df['是否追踪到'] != '转出外省') & (df['地市'] != '广州市')

    #推送外区
    hexinmijie_waiqu = hexinmijie_active & (df['是否追踪到'] != '转出外省') & (df['地市'] == '广州市') & (df['区县'] != '白云区')

    #区外推送
    hexinmijie_quwai = hexinmijie_leiji & (df['创建单位'] != '白云区疾病预防控制中心') & (df['是否核心密接'] == '是')

    #我区应管核心密接
    hexinmijie_yingguan = hexinmijie_leiji & (df['地市'] == '广州市') & (df['区县'] == '白云区') & (df['是否追踪到'] != '转出外省')

    #日期类型修改
    df['录入时间'] = pd.to_datetime(df['录入时间']).dt.date
    # print(df['审核时间'])
    df['审核时间'] = pd.to_datetime(df['审核时间']).dt.date
    df['结束时间'].fillna('2037-12-31',inplace=True)
    df['结束时间'] = pd.to_datetime(df['结束时间']).dt.date
    #开始时间
    starttime = pd.to_datetime(df[(df['是否排除密接/次密'] != '是') & name_type]['录入时间'].min()).date()
    # print('该病例开始时间',starttime)

    #结束时间

    if len(df[(df['是否排除密接/次密'] != '是') & name_type & (df['转归'] == '继续观察')]) ==0:
        endtime = pd.to_datetime(df[mijie_leiji & (df['转归'] == '解除观察')]['结束时间'].max()).date()
        # print('该病例结束时间', endtime)
    else:
        endtime = datetime.datetime.now().date()
        # print('病例还未解除')
        # endtime = datetime.date(2022,11,7)

    # print(df[hexinmijie_yingguan & (df['录入时间'] <= endtime)]['审核时间'])

    for i in range((endtime-starttime).days+1):
        day = starttime + datetime.timedelta(days=i)
        # print(day)
        # 当天时间段包含的数据
        time_luru = (starttime <= df['录入时间']) & (df['录入时间'] <=day)
        #############核心密接###############
        # 当天时间累计甄别核心密接
        hexinmijie_leiji_total = len(df[hexinmijie_leiji & time_luru])
        # 当天时间主动甄别核心密接
        hexinmijie_active_total = len(df[hexinmijie_active & time_luru])
        # 当天时间推送外区核心密接
        hexinmijie_waiqu_total = len(df[hexinmijie_waiqu & time_luru])
        # 当天时间推送外市核心密接
        hexinmijie_outcity_total = len(df[hexinmijie_outcity & time_luru])
        # 当天时间推送外省核心密接
        hexinmijie_pro_total = len(df[hexinmijie_pro & time_luru])
        # 当天区外推送核心密接
        hexinmijie_quwai_total = len(df[hexinmijie_quwai & time_luru])
        # 当天我区应管核心密接
        hexinmijie_yingguan_total = len(df[hexinmijie_yingguan & time_luru])
        # 当天已落地核心密接
        hexinmijie_luodi = len(df[hexinmijie_yingguan & time_luru & (df['审核时间'] <= day)])
        # 当天在管核心密接
        # df[mijie_yingguan & time_luru & (df['审核时间'] <= day) & ((df['结束时间'] > day) | (df['结束时间'] == np.nan))].to_excel('在管密接.xlsx')
        hexinmijie_zaiguan = len(df[hexinmijie_yingguan & time_luru & (df['审核时间'] <= day) & (df['结束时间'] > day)])
        # mijie_zaiguan = len(df[mijie_yingguan & time_luru & (df['审核时间'] <= day) & (df['结束时间'] > day)]) + len(df[mijie_yingguan & time_luru & (df['审核时间'] <= day) & (df['结束时间'] == np.nan)])
        # mijie_zaiguan = len(df[mijie_yingguan & time_luru & (df['审核时间'] <= day) & (df['结束时间'] == '空白')])

        ###管控情况
        #核心密接集中
        hexinmijie_jizhong = len(df[hexinmijie_yingguan & time_luru & (df['审核时间'] <= day) & (df['结束时间'] > day) & (df['医学观察方式'] == '集中')])
        # 核心密接待转运
        hexinmijie_daizhaunyun = len(df[hexinmijie_yingguan & time_luru & (df['审核时间'] <= day) & (df['结束时间'] > day) & (df['医学观察方式'] == '待转运')])
        # 核心密接居家
        hexinmijie_jujia = len(df[hexinmijie_yingguan & time_luru & (df['审核时间'] <= day) & (df['结束时间'] > day) & (df['医学观察方式'] == '居家')])
        # 核心密接医院隔离
        hexinmijie_yiyuan = len(df[hexinmijie_yingguan & time_luru & (df['审核时间'] <= day) & (df['结束时间'] > day) & (df['医学观察方式'] == '医院隔离')])

        # 当天接触隔离核心密接
        hexinmijie_jiechu = len(df[hexinmijie_yingguan & time_luru & (df['审核时间'] <= day) & (df['结束时间'] <= day) & (df['转归'] == '解除观察')])
        # 当天转为确诊核心密接
        # mijie_quezhen = len(df[mijie_yingguan & time_luru & (df['审核时间'] <= day) & (df['转归'] == '转为确诊') & (df['转归'] == '转为阳性')])
        hexinmijie_quezhen = len(df[hexinmijie_yingguan & time_luru & (df['审核时间'] <= day) & (df['转归'] == '转为确诊')]) + len(df[hexinmijie_yingguan & time_luru & (df['审核时间'] <= day) & (df['转归'] == '转为阳性')])
        # 当天正在核实追踪核心密接
        # hexinmijie_zhuizong = len(df[hexinmijie_yingguan & time_luru & (df['审核时间'] > day)]) + len(df[hexinmijie_yingguan & time_luru & (np.isnan(df['审核时间']))])
        hexinmijie_zhuizong = len(df[hexinmijie_yingguan & time_luru & (df['审核时间'] > day)])
        # 当天核心密接新增人数
        hexinmijie_xinzhen = len(df[hexinmijie_yingguan & (df['录入时间'] == day)])







        #############密接###############
        #当天时间累计甄别密接
        mijie_leiji_total = len(df[mijie_leiji & time_luru])
        # 当天时间主动甄别密接
        mijie_active_total = len(df[mijie_active & time_luru])
        # 当天时间推送外区密接
        mijie_waiqu_total = len(df[mijie_waiqu & time_luru])
        #当天时间推送外市密接
        mijie_outcity_total = len(df[mijie_outcity & time_luru])
        #当天时间推送外省密接
        mijie_pro_total = len(df[mijie_pro & time_luru])
        #当天区外推送密接
        mijie_quwai_total = len(df[mijie_quwai & time_luru])
        #当天我区应管密接
        mijie_yingguan_total = len(df[mijie_yingguan & time_luru])
        #当天已落地密接
        mijie_luodi = len(df[mijie_yingguan & time_luru & (df['审核时间'] <= day)])
        # 当天在管密接
        # df[mijie_yingguan & time_luru & (df['审核时间'] <= day) & ((df['结束时间'] > day) | (df['结束时间'] == np.nan))].to_excel('在管密接.xlsx')
        mijie_zaiguan = len(df[mijie_yingguan & time_luru & (df['审核时间'] <= day) & (df['结束时间'] > day)])
        # mijie_zaiguan = len(df[mijie_yingguan & time_luru & (df['审核时间'] <= day) & (df['结束时间'] > day)]) + len(df[mijie_yingguan & time_luru & (df['审核时间'] <= day) & (df['结束时间'] == np.nan)])
        # mijie_zaiguan = len(df[mijie_yingguan & time_luru & (df['审核时间'] <= day) & (df['结束时间'] == '空白')])

        ###管控情况
        #密接集中
        mijie_jizhong = len(df[mijie_yingguan & time_luru & (df['审核时间'] <= day) & (df['结束时间'] > day) & (df['医学观察方式'] == '集中')])
        #密接待转运
        mijie_daizhaunyun = len(df[mijie_yingguan & time_luru & (df['审核时间'] <= day) & (df['结束时间'] > day) & (df['医学观察方式'] == '待转运')])
        #密接居家
        mijie_jujia = len(df[mijie_yingguan & time_luru & (df['审核时间'] <= day) & (df['结束时间'] > day) & (df['医学观察方式'] == '居家')])
        #密接医院隔离
        mijie_yiyuan = len(df[mijie_yingguan & time_luru & (df['审核时间'] <= day) & (df['结束时间'] > day) & (df['医学观察方式'] == '医院隔离')])
        # 当天接触隔离密接
        mijie_jiechu = len(df[mijie_yingguan & time_luru & (df['审核时间'] <= day) & (df['结束时间'] <= day) & (df['转归'] == '解除观察')])
        # 当天转为确诊密接
        # mijie_quezhen = len(df[mijie_yingguan & time_luru & (df['审核时间'] <= day) & (df['转归'] == '转为确诊') & (df['转归'] == '转为阳性')])
        mijie_quezhen = len(df[mijie_yingguan & time_luru & (df['审核时间'] <= day) & (df['转归'] == '转为确诊')]) + len(df[mijie_yingguan & time_luru & (df['审核时间'] <= day) & (df['转归'] == '转为阳性')])
        # 当天正在核实追踪密接
        # mijie_zhuizong = len(df[mijie_yingguan & time_luru & (df['审核时间'] > day)]) + len(df[mijie_yingguan & time_luru & (np.isnan(df['审核时间']))])
        mijie_zhuizong = len(df[mijie_yingguan & time_luru & (df['审核时间'] > day)])

        #当天密接新增人数
        mijie_xinzhen = len(df[mijie_yingguan & (df['录入时间'] == day)])


        df_new = [day,'汇总',hexinmijie_leiji_total,hexinmijie_active_total,hexinmijie_waiqu_total,hexinmijie_outcity_total,hexinmijie_pro_total,hexinmijie_quwai_total,
                   hexinmijie_yingguan_total,hexinmijie_luodi,hexinmijie_zaiguan,hexinmijie_jizhong,hexinmijie_daizhaunyun,hexinmijie_jujia,hexinmijie_yiyuan,hexinmijie_jiechu,hexinmijie_quezhen,hexinmijie_zhuizong,hexinmijie_xinzhen,
                  mijie_leiji_total,mijie_active_total,mijie_waiqu_total,mijie_outcity_total,mijie_pro_total,mijie_quwai_total,
                   mijie_yingguan_total,mijie_luodi,mijie_zaiguan,mijie_jizhong,mijie_daizhaunyun,mijie_jujia,mijie_yiyuan,mijie_jiechu,mijie_quezhen,mijie_zhuizong,mijie_xinzhen
                   ]
        data_list.append(df_new)




    # print(data_list)
    data_new = pd.DataFrame(data_list,
    columns=['开始时间','病例（专项名称）','累计甄别核心密接','主动甄别核心密接','推送外区核心密接','推送外市核心密接','推送外省核心密接','区外推送核心密接','我区应管核心密接','核心密接已落地','核心密接在管','核心密接集中','核心密接待转运','核心密接居家','核心密接医院隔离','核心密接解除隔离','核心密接转为确诊','核心密接核实追踪','核心密接新增人数',
    '累计甄别密接','主动甄别密接','推送外区密接','推送外市密接','推送外省密接','区外推送密接','我区应管密接','密接已落地','密接在管','密接集中','密接待转运','密接居家','密接医院隔离','密接解除隔离','密接转为确诊','密接核实追踪','密接新增人数'])
    data_new.set_index('开始时间', inplace=True)
    # print(data_new)
    df_old = pd.read_excel(rf'./output/{excel_name}.xlsx')
    row_old = df_old.shape[0]
    book = load_workbook(rf'./output/{excel_name}.xlsx')
    writer2 = pd.ExcelWriter(rf'./output/{excel_name}.xlsx', engine='openpyxl')
    writer2.book = book
    try:
        writer2.sheets = dict((ws.title, ws) for ws in book.worksheets)
    except:
        writer2._sheets = dict((ws.title, ws) for ws in book.worksheets)
    data_new.to_excel(writer2, sheet_name='数据',startrow=row_old + 1, index='开始时间', header=False)
    writer2.save()


def Chart():

    df = pd.read_excel(rf'./output/{excel_name}.xlsx')
    df_old_row = df.shape[0]
    wb = load_workbook(rf'./output/{excel_name}.xlsx')
    wb_sheet = wb['数据']
    ws = wb.create_sheet()



    '''密接次密涉我区人数'''
    chart1 = BarChart()
    chart1.type = "col"  # 纵向柱形图
    chart1.style = 11
    chart1.title = "核心密接、密接涉我区人数"  # 图表标题
    chart1.y_axis.title = '人数'  # 纵坐标轴标题
    chart1.x_axis.title = '下库时间'  # 横坐标轴标题
    data = Reference(wb_sheet, min_col=9,  min_row=1,max_col=9, max_row=df_old_row+1)  # 核心密接
    data2 = Reference(wb_sheet, min_col=26,  min_row=1,max_col=26, max_row=df_old_row+1) #密接
    cats = Reference(wb_sheet, min_col=1, min_row=1,max_col=1,max_row=df_old_row+1)  # 类别对象
    chart1.add_data(data, titles_from_data=True)
    chart1.add_data(data2, titles_from_data=True)
    chart1.set_categories(cats)  # 传入类别范围
    chart1.shape = 20
    chart1.varyColors = 'red'
    ws.add_chart(chart1, "A10")
    #显示数据标签
    chart1.dataLabels = DataLabelList()
    chart1.dataLabels.showVal = True
    chart1.dataLabels.position = 'ctr'
    # wb.save('test.xlsx')



    '''密接次密在管人员情况'''
    chart2 = BarChart()
    chart2.type = "col"  # 纵向柱形图
    chart2.style = 10
    chart2.title = "在管人员情况"  # 图表标题
    chart2.y_axis.title = '人数'  # 纵坐标轴标题
    chart2.x_axis.title = '下库时间'  # 横坐标轴标题
    data_zaiguan = Reference(wb_sheet, min_col=12,  min_row=1,max_col=15, max_row=df_old_row+1)  # 数据引用范围，
    data2_zaiguan = Reference(wb_sheet, min_col=29,  min_row=1,max_col=32, max_row=df_old_row+1) #次密
    chart2.add_data(data_zaiguan, titles_from_data=True)
    chart2.set_categories(cats)  # 传入类别范围
    chart2.shape = 4
    chart2.grouping = 'stacked'
    chart2.overlap = 100
    ws.add_chart(chart2,"A30")
    ##
    chart2_cimi = BarChart()
    chart2_cimi.type = "col"  # 纵向柱形图
    chart2_cimi.style = 14
    chart2_cimi.title = "在管人员情况"  # 图表标题
    chart2_cimi.y_axis.title = '人数'  # 纵坐标轴标题
    chart2_cimi.x_axis.title = '下库时间'  # 横坐标轴标题
    chart2_cimi.add_data(data2_zaiguan, titles_from_data=True)
    chart2_cimi.set_categories(cats)  # 传入类别范围
    chart2_cimi.shape = 4
    chart2_cimi.shape = 4
    chart2_cimi.grouping = 'stacked'
    chart2_cimi.overlap = 100
    ws.add_chart(chart2_cimi,"A50")
    #显示数据标签
    chart2.dataLabels = DataLabelList()
    chart2.dataLabels.showVal = True
    chart2.dataLabels.position = 'ctr'
    chart2_cimi.dataLabels = DataLabelList()
    chart2_cimi.dataLabels.showVal = True
    chart2_cimi.dataLabels.position = 'ctr'
    # wb.save('test.xlsx')



    '''密接次密涉我区新增人数'''
    chart3 = LineChart()
    chart3.type = "Line Chart"  # 纵向柱形图
    chart3.style = 14
    chart3.title = "密接次密涉我区新增人数"  # 图表标题
    chart3.y_axis.title = '人数'  # 纵坐标轴标题
    chart3.x_axis.title = '下库时间'  # 横坐标轴标题
    data_xinzeng = Reference(wb_sheet, min_col=19,  min_row=1,max_col=19, max_row=df_old_row+1)  # 数据引用范围，
    data2_xinzeng = Reference(wb_sheet, min_col=36,  min_row=1,max_col=36, max_row=df_old_row+1) #次密
    chart3.add_data(data_xinzeng, titles_from_data=True)
    chart3.add_data(data2_xinzeng, titles_from_data=True)
    chart3.set_categories(cats)  # 传入类别范围
    chart3.shape = 4
    ws.add_chart(chart3, "A70")
    #显示数据标签
    # chart3.dataLabels = DataLabelList()
    # chart3.dataLabels.showVal = True
    # chart3.dataLabels.position = 't'
    wb.save(rf'./output/{excel_name}.xlsx')


def Chart2():
    data_chart = report_newusual()
    import re
    import test

    import pandas as pd
    from openpyxl import load_workbook
    import openpyxl
    from openpyxl.drawing.image import Image
    import os

    df = pd.read_excel(rf'./output/{excel_name}.xlsx')

    num = len(df['开始时间'])
    Time = []
    for i in df['开始时间'].dt.date:
        Time.append(str(i))

    hexinmijie_workable = []
    for i in df['我区应管核心密接']:
        hexinmijie_workable.append(i)

    mijie_workable = []
    for i in df['我区应管密接']:
        mijie_workable.append(i)

    # 非核心密接
    workable = []
    for i in range(num):
        x = mijie_workable[i] - hexinmijie_workable[i]
        workable.append(x)


    hexinmijie_new = []
    for i in df['核心密接新增人数']:
        hexinmijie_new.append(i)

    mijie_new = []
    for i in df['密接新增人数']:
        mijie_new.append(i)




    html =f'''
    <!DOCTYPE html>
    <html>
      <head>
        <meta charset="utf-8" />
        <title>Echart</title>
        <script src=".\echarts.min.js"></script>
      </head>
      <body>
        <div id="main" style="width: 800px;height:800px;"></div>
        <div id="main2" style="width: 800px;height:800px;"></div>
        <div id="main3" style="width: 800px;height:800px;"></div>
        <div id="main4" style="width: 800px;height:800px;"></div>
        <script type="text/javascript">
        var canvas = document.getElementsByTagName("canvas");
         var chartDom = document.getElementById('main');
         var myChart = echarts.init(chartDom);
         var chartDom2 = document.getElementById('main2');
         var myChart2 = echarts.init(chartDom2);
         var chartDom3 = document.getElementById('main3');
         var myChart3 = echarts.init(chartDom3);
         var chartDom4 = document.getElementById('main4');
         var myChart4 = echarts.init(chartDom4);
         var option;
         var option2;
         var option3;
         var option4;
    option = {{
      legend: {{
        data: ['我区应管核心密接','我区应管密接','核心密接新增人数','密接新增人数']
      }},
      xAxis: [
        {{
          type: 'category',
          interval:0,
          data: {Time},
          axisPointer: {{
            type: 'shadow',
            splitnumber:{num}
          }},
          axisLabel: {{
                rotate:20
                    }}
        }}
      ],
      yAxis: [
        {{
          type: 'value',
          name: '涉我区核密、密接',
          axisLabel: {{
            formatter: '{{value}} 人'
          }}
        }},
        {{
          type: 'value',
          name: '新增人数',
          min:0,
          max:500,
          interval:50,
          axisLabel: {{
            formatter: '{{value}} 人'
          }}
        }},
      ],
      series: [
        {{
          name: '我区应管核心密接',
          type: 'bar',
          data: {hexinmijie_workable},
          stack:'涉我区密接',
          label:{{
          show:true,
          position:'top'
          }}
        }},
        {{
            name: '我区应管密接',
            type: 'bar',
            stack: '涉我区密接',
            label: {{
                normal: {{
                    show: true,   
                    position: 'top',
                    formatter: function (params) {{
                        return params.value + {hexinmijie_workable}[params.dataIndex]
                    }},
                    textStyle: {{color: '#000'}} 
                }}
            }},
            itemStyle: {{
                normal: {{
                    //color: '#999999'    //设置柱状图颜色 
                }}
            }},
            data: {workable}
        }},

        {{
          name: '核心密接新增人数',
          type: 'line',
          yAxisIndex:1,
          data: {hexinmijie_new},
          label:{{
          show:true,
          position:'top',
          }}
        }},
        {{
          name: '密接新增人数',
          type: 'line',
          data: {mijie_new},
          label:{{
          show:true,
          position:'bottom'
          }}
        }},
      ]
    }};
    option && myChart.setOption(option);
    
    option2 = {{
  legend: {{
    top: 'top'
  }},
  toolbox: {{
    show: true,
    feature: {{
      mark: {{ show: true }},
      dataView: {{ show: true, readOnly: false }},
      restore: {{ show: true }},
      saveAsImage: {{ show: true }}
    }}
  }},
  series: [
    {{
      name: '涉我区核心密接管控情况',
      type: 'pie',
      radius: [50, 250],
      center: ['50%', '50%'],
      roseType: 'area',
      itemStyle: {{
        borderRadius: 8
      }},
      data: [
        {{ value: {data_chart['核心密接核实追踪']}, name: '核心密接核实追踪' }},
        {{ value: {data_chart['核心密接待转运']}, name: '核心密接待转运' }},
        {{ value: {data_chart['核心密接集中']}, name: '核心密接集中' }},
        {{ value: {data_chart['核心密接居家']}, name: '核心密接居家' }},
        {{ value: {data_chart['核心密接医院隔离']}, name: '核心密接医院隔离' }},
        {{ value: {data_chart['核心密接转为确诊']}, name: '核心密接转为确诊' }},
        {{ value: {data_chart['核心密接解除观察']}, name: '核心密接解除观察' }}
      ]
    }}
  ]
}};
option2 && myChart2.setOption(option2);

    option3 = {{
  legend: {{
    top: 'top'
  }},
  toolbox: {{
    show: true,
    feature: {{
      mark: {{ show: true }},
      dataView: {{ show: true, readOnly: false }},
      restore: {{ show: true }},
      saveAsImage: {{ show: true }}
    }}
  }},
  series: [
    {{
      name: '涉我区密接管控情况',
      type: 'pie',
      radius: [50, 250],
      center: ['50%', '50%'],
      roseType: 'area',
      itemStyle: {{
        borderRadius: 8
      }},
      data: [
        {{ value: {data_chart['密接核实追踪']}, name: '密接核实追踪' }},
        {{ value: {data_chart['密接待转运']}, name: '密接待转运' }},
        {{ value: {data_chart['密接集中']}, name: '密接集中' }},
        {{ value: {data_chart['密接居家']}, name: '密接居家' }},
        {{ value: {data_chart['密接医院隔离']}, name: '密接医院隔离' }},
        {{ value: {data_chart['密接转为确诊']}, name: '密接转为确诊' }},
        {{ value: {data_chart['密接解除观察']}, name: '密接解除观察' }}
      ]
    }}
  ]
}};
option3 && myChart3.setOption(option3);


option4 = {{
  legend: {{
    top:'top'
  }},
  series: [
    {{
      name: 'Access From',
      type: 'pie',
      radius: ['40%', '70%'],
      avoidLabelOverlap: false,
      itemStyle: {{
        borderRadius: 10,
        borderColor: '#fff',
        borderWidth: 2
      }},
      label:{{
            show:true,
            formatter:'{{d}}人'
      }},
      labelLine: {{
        show: true,
      }},
      data: [
        {{ value: {data_chart['密接核实追踪']}, name: '密接核实追踪' }},
        {{ value: {data_chart['密接待转运']}, name: '密接待转运' }},
        {{ value: {data_chart['密接集中']}, name: '密接集中' }},
        {{ value: {data_chart['密接居家']}, name: '密接居家' }},
        {{ value: {data_chart['密接医院隔离']}, name: '密接医院隔离' }},
        {{ value: {data_chart['密接转为确诊']}, name: '密接转为确诊' }},
        {{ value: {data_chart['密接解除观察']}, name: '密接解除观察' }}
      ]
    }}
  ]
}};

option4 && myChart4.setOption(option4);
        </script>
      </body>
    </html>
    '''

    with open('.\html文件夹\我区应管核心密接、密接可视化.html', 'w', encoding='utf-8') as f:
        f.write(html)

    # with open('.\html文件夹\我区应管核心密接可视化.html', 'w', encoding='utf-8') as f:
    #     f.write(html2)

    # time.sleep(2)
    # os.system('.\html文件夹\我区应管密接次密可视化图表.html')
    # time.sleep(2)

    # wb = load_workbook(rf'D:\同步空间\专项报表\特提1015-1025关联201人汇总.xlsx')
    # names = list(wb.sheetnames)
    # if '我区应管密接次密可视化图表' in names:
    #     del wb['我区应管密接次密可视化图表']
    #
    # ws = wb.create_sheet('我区应管密接次密可视化图表')
    # sheet = wb['我区应管密接次密可视化图表']
    # img = Image(r"C:\Users\Administrator\Downloads\我区应管密接次密可视化图表.png")
    # sheet.add_image(img,'A2')
    # wb.save(rf'D:\同步空间\专项报表\特提1015-1025关联201人汇总.xlsx')
    # time.sleep(2)
    # os.remove(r"C:\Users\Administrator\Downloads\我区应管密接次密可视化图表.png")




# dealwith()
# statistics()
# report_newusual()
# Chart()
Chart2()






