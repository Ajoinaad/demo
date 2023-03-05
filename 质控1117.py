# !/usr/bin/env python
# coding: utf-8
import shutil
import os
import time
import datetime
import warnings
from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
import pandas as pd
from openpyxl import load_workbook
import numpy as np
import openpyxl
from docx import Document
from docx.oxml.ns import qn
from openpyxl.utils import get_column_letter
now_time = datetime.datetime.now()
warnings.filterwarnings('ignore')
name = ' '  # 所有病例数据处理
title = '抬头' # 默认抬头

# 保存报表数据
data_table = []
Statistics = 0


# 先计算密接类型
class Original:
    report = pd.read_excel('批量汇报.xlsx', sheet_name='汇报专项')
    doc = Document()

    # 总数
    def __init__(self, name='病例名称'):
        self.name = name

    # 1.文件名
    def open_file(self):
        # print('-' * 80)
        files = r"input"  # 打开文件夹
        lists = os.listdir(files)  # 列出目录的下所有文件和文件夹保存到lists
        lists.sort(key=lambda fn: os.path.getmtime(files + "\\" + fn))  # 按时间排序
        file_new = os.path.join(files, lists[-1])
        return file_new

    # 2.数据处理
    def dealwith(self):
        print('\n\n\n')
        print('数据正在处理中...')
        # 0.删除旧文件,创建新文件
        try:
            shutil.rmtree(rf'D:\工作源文件')
        except:
            print('数据处理出现异常,请关闭所有文件夹/文件,并检查D:\盘是否有D:\工作源文件,有则删除,没有请重新运行')

        os.makedirs(rf'D:\工作源文件')
        # 1.读取最新需要处理文件
        df_deal = pd.read_excel(f'./{Original().open_file()}')
        # 1.1重复值处理(目前还不确定是否要弄)
        # df_deal = df.drop_duplicates(subset=['有效证件号','是否排除密接/次密','密接类型','转归'], keep='first', inplace=True)
        # 2.数据空值处理
        df_deal['审核时间'].fillna('空白', inplace=True)
        df_deal['镇（街道）'].fillna('空白', inplace=True)
        df_deal['转出目的省(直辖市)'].fillna('不明', inplace=True)
        df_deal['转归'].fillna('空白', inplace=True)
        df_deal['医学观察场所名称'].fillna('空白', inplace=True)
        # 2.2日期处理
        df_deal['一码通入住日期'].fillna('1970-12-30 23:59', inplace=True)
        df_deal['最后接触日期'].fillna('1970-12-31 23:59', inplace=True)
        df_deal['一码通入住日期'] = pd.to_datetime(df_deal['一码通入住日期']).dt.floor('d')
        df_deal['最后接触日期'] = pd.to_datetime(df_deal['最后接触日期']).dt.floor('d')
        df_deal['应解除观察日期'] = pd.to_datetime(df_deal['应解除观察日期']).dt.floor('d')
        # 2.3数据空值预处理
        df_deal['审核时间'].fillna('空白', inplace=True)
        df_deal['镇（街道）'].fillna('空白', inplace=True)

        # 375个场所
        df_deal = df_deal.get(
            ['地市', '区县', '镇（街道）', 'ID', '姓名', '国籍', '性别', '年龄', '有效证件号', '联系方式', '目前所处位置', '现住址', '职业', '工作单位',
             '密接/次密发现途径', '是否核心密接', '是否排除密接/次密', '关联病例','是否境外输入病例', '关联密接', '密接类型', '与患者关系', '接触地点', '最后接触日期', '应解除观察日期', '关联重点场所',
             '转归', '备注', '是否追踪到', '审核时间', '录入时间','一码通入住日期','一码通入住酒店', '医学观察方式', '医学观察场所名称','创建单位', '转出目的省(直辖市)', '居家隔离原因'])
        df_deal.to_csv(f'D:/工作源文件/{Original().open_file()[5:-5]}.csv', index=False)  # 保存最新版数据

    def report_roster(self):
        # 先计算密接类型
        sick_type = '密接'
        # # 读取处理文件
        df = pd.read_csv(f'D:/工作源文件/{Original().open_file()[5:-5]}.csv')
        # 名字类别
        name_List = name.split()
        name_type = df['关联病例'].str.contains('|'.join(name_List))
        print('程序正在进行数据筛选，筛选完成后才会生成名单.....')
        # 密接类型
        type = ['密切接触者']
        for i in range(0, len(type)):




            '''筛选数据'''
            count = (df['是否排除密接/次密'] != '是') & (df['密接类型'] == type[i]) & name_type
            # 我区主动甄别
            active = count & (df['创建单位'] == '白云区疾病预防控制中心')
            # 推送外省管控人数
            province = active & (df['是否追踪到'] == '转出外省')
            # 推送外市管控人数
            city = active & (df['是否追踪到'] != '转出外省') & (df['地市'] != '广州市')
            # 推送外区管控人数
            area = active & (df['是否追踪到'] != '转出外省') & (df['地市'] == '广州市') & (df['区县'] != '白云区')
            # 归属白云区
            local = active & (df['是否追踪到'] != '转出外省') & (df['地市'] == '广州市') & (df['区县'] == '白云区')
            # 外区推送我区,我区推送外省(求Y的值) :区外推送X人（推送外省'Y'人）
            out = count & (df['创建单位'] != '白云区疾病预防控制中心') & (df["是否追踪到"] == "转出外省")
            # 第二点
            # 涉及我区应管
            must = count & (df['是否追踪到'] != '转出外省') & (df['地市'] == '广州市') & (df['区县'] == '白云区') & (df['是否境外输入病例'] == '否') & (df['创建单位'] == '白云区疾病预防控制中心')
            # 已落地
            workable = (must & (df['审核时间'] != '空白')) + (must & (df['转归'] == '转为确诊'))
            # 目前在管
            atpresent = workable & (df['转归'].str.contains('继续'))
            # 集中
            focus = atpresent & (df['医学观察方式'].str.contains('集中'))
            # 待转运
            wait = atpresent & (df['医学观察方式'].str.contains('待转运'))
            wait_count = len(df[wait])
            # 居家
            home = atpresent & (df['医学观察方式'].str.contains('居家'))
            # 医院隔离
            hospital = atpresent & (df['医学观察方式'].str.contains('医院'))
            # 解除隔离
            remove = must & (df['转归'].str.contains('解除')) & (df['审核时间'] != '空白')
            # 正在核实追踪
            track = must & (df['审核时间'] == '空白') & (df['转归'] != '转为确诊')
            # 转为确诊
            sun = must & (df['转归'] == '转为确诊')



            # 质控名单
            # 1、镇（街道）
            road_List = ['同德', '松洲', '黄石', '石井', '鹤龙', '龙归', '金沙', '嘉禾', '云城', '三元里', '京溪', '同和', '人和', '均禾', '大源',
                         '太和', '白云湖',
                         '景泰', '棠景', '永平', '江高', '石门', '新市', '钟落潭']
            wrong_must = must & (df['镇（街道）'].str.contains('|'.join(road_List)) == False) & (df['医学观察方式'] != '居家') & (df['医学观察方式'] != '集中')
            if (len(df[wrong_must])) != 0:
                df[wrong_must].to_excel(rf'./质控名单/1.{sick_type}镇街维护.xlsx', index=False)
                print(f'已生成{sick_type}街道维护名单')

            # # 2、目前所处位置
            # wrong_position = must & (((df['目前所处位置'].str.contains('省')) & (df['目前所处位置'].str.contains('市')) & (df['目前所处位置'].str.contains('区')) & (df['目前所处位置'].str.contains('街|镇|巷')))== False)
            # if (len(df[wrong_position])) != 0:
            #     df[wrong_position].to_excel(rf'./质控名单/2.{sick_type}目前所处位置维护.xlsx', index=False)
            #     print(f'已生成{sick_type}目前所处位置维护名单')
            #
            # # 3、现住址
            # wrong_live = must & (((df['现住址'].str.contains('省')) & (df['现住址'].str.contains('市')) & (
            # df['现住址'].str.contains('区')) & (df['现住址'].str.contains('街|镇|路'))) == False)
            # if (len(df[wrong_live])) != 0:
            #     df[wrong_live].to_excel(rf'./质控名单/3.{sick_type}现住址维护.xlsx', index=False)
            #     print(f'已生成{sick_type}现住址维护名单')
            #
            # # 4、职业
            # job_list = ['幼托儿童','散居儿童','学生','教师','保育员及保姆','餐饮食品业','商业服务','工人','民工','牧民','渔(船)民','干部职员','离退人员','家务待业','医护人员','不详','其他']
            # check_job = must & (df['职业'].str.contains('|'.join(job_list)) == False)
            # if (len(df[check_job])) != 0:
            #     df[check_job].to_excel(rf'./质控名单/4.{sick_type}职业维护.xlsx', index=False)
            #     print(f'{sick_type}职业维护名单已生成')
            #
            # # 5、工作单位
            # work_units_list = ['医护人员' ,'学生' ,'老师']
            # check_work_units = must & df['职业'].str.contains('|'.join(work_units_list)) & (df['工作单位'].isnull())
            # if (len(df[check_work_units])) != 0:
            #     df[check_work_units].to_excel(rf'./质控名单/5.{sick_type}特定职业工作单位为空维护.xlsx', index=False)
            #     print(f'{sick_type}特定职业工作单位为空维护名单已生成')
            #
            # # 6、密接/次密发现途径
            # df_find = must & (df['密接/次密发现途径'].isnull())
            # if (len(df[ df_find])) != 0:
            #     df[df_find].to_excel(rf'./质控名单/6.{sick_type}发现途径为空维护.xlsx', index=False)
            #     print(f'{sick_type}发现途径为空名单已生成')
            #
            # # 8、应解除观察日期
            # way_list = ['继续观察', '解除观察', '转为阳性', '转为确诊']
            # wrong_way = must & (df['转归'].str.contains('|'.join(way_list)) == False)
            # if (len(df[wrong_way])) != 0:
            #     df[wrong_way].to_excel(rf'./质控名单/8.{sick_type}转归维护.xlsx', index=False)
            #     print(f'{sick_type}转归维护名单已生成')
            # wrong_data = must & ((pd.to_datetime(df['应解除观察日期'])-pd.to_datetime(df['最后接触日期'])).map(lambda x: x/np.timedelta64(1,'D')) > 5)
            # df[wrong_data].to_excel(rf'./质控名单/8.{sick_type}应解除观察日期维护.xlsx',index=False)
            # print(f'{sick_type}应解除观察日期维护名单已生成')

            # 9、是否追踪到
            track = (must & (df['是否追踪到'].isnull())) | (must & (df['是否追踪到'] == '否'))
            if (len(df[track])) != 0:
                df[track].to_excel(
                rf'./质控名单/9.{sick_type}是否追踪到为空的维护.xlsx',
                index=False)
                print(f'{sick_type}是否追踪到为空的维护名单已生成')

            # 10、转出目的省（直辖市）
            wrong_turn_out = ['河北','山西','辽宁','吉林','黑龙江','江苏','浙江','安徽','福建','江西','山东','河南','湖北','湖南','海南','四川','贵州','云南','陕西','甘肃','青海','台湾','内蒙','广西','西藏','宁夏','新疆','北京','天津','上海','重庆','澳门','香港','境外']
            turn_out = province & (((df['转出目的省(直辖市)'] == '不明')) | (df['转出目的省(直辖市)'].str.contains('|'.join(wrong_turn_out)) == False))
            if (len(df[turn_out])) != 0:
                df[turn_out].to_excel(
                rf'./质控名单/10.{sick_type}转出目的省(直辖市)为空的维护.xlsx',
                index=False)
                print(f'{sick_type}转出目的省(直辖市)为空的维护名单已生成')
            #
            # # 11、地市/区县
            # place = ['深圳市','广州市','珠海市','东莞市','佛山市','中山市','惠州市','汕头市','江门市','湛江市','肇庆市','梅州市','茂名市','阳江市','清远市','韶关市','揭阳市','汕尾市','潮州市','河源市','云浮市']
            # wrong_place = count & (df['是否追踪到'] != '转出外省') & (df['地市'].str.contains('|'.join(place)) ==False)
            # if (len(df[wrong_place])) != 0:
            #     df[ wrong_place].to_excel(
            #         rf'./质控名单/11.{sick_type}地市的维护.xlsx',
            #         index=False)
            #     print(f'{sick_type}地市的维护名单已生成')
            # place_qu = ['荔湾区','越秀区','海珠区','天河区','白云区','黄埔区','番禺区','花都区','南沙区','从化区','增城区']
            # must1 = count & (df['是否追踪到'] != '转出外省') & (df['地市'] == '广州市')
            # wrong_place_qu = must1 & (df['区县'].str.contains('|'.join(place_qu)) == False)
            # if (len(df[wrong_place_qu])) != 0:
            #     df[wrong_place_qu].to_excel(
            #     rf'./质控名单/11.{sick_type}区县的维护.xlsx',
            #     index=False)
            #     print(f'{sick_type}区县的维护名单已生成')

            # 12、医学观察方式、医学观察场所名称、一码通入住日期
            # 【一码通入住日期】和【医学观察场所】“空白字段”归纳为待转运
            if sick_type == '密接':
                change_dzy = (must & (df['转归'] == '继续观察') & (df['审核时间'] != '空白') & (df['医学观察方式'] == '集中') & (df['医学观察场所名称'] == '空白') & (df['一码通入住日期'] == '1970-12-30')) | \
                             (must & (df['转归'] == '继续观察') & (df['审核时间'] != '空白') & (df['医学观察方式'] == '居家') & df['居家隔离原因'].isnull())
                if (len(df[change_dzy])) != 0:
                    df[change_dzy].to_excel(
                        rf'./质控名单/12.{sick_type}医学观察方式（改待转运）的维护.xlsx',index=False)
                    print(f'{sick_type}医学观察方式（改待转运）的维护名单已生成')


                change_jz = (must & (df['转归'] == '继续观察') & (df['审核时间'] != '空白') & (df['医学观察方式'] == '待转运') & ((((pd.to_datetime(df['一码通入住日期'])-pd.to_datetime(df['最后接触日期'])).map(lambda x: x/np.timedelta64(1,'D')) < 5) & ((pd.to_datetime(df['一码通入住日期'])-pd.to_datetime(df['最后接触日期'])).map(lambda x: x/np.timedelta64(1,'D')) >= 0)) & (df['一码通入住日期'] != '1970-12-30'))) \
                            | (must & (df['转归'] == '继续观察') & (df['审核时间'] != '空白') & (df['医学观察方式'] == '居家') & ((((pd.to_datetime(df['一码通入住日期'])-pd.to_datetime(df['最后接触日期'])).map(lambda x: x/np.timedelta64(1,'D')) < 5) & ((pd.to_datetime(df['一码通入住日期'])-pd.to_datetime(df['最后接触日期'])).map(lambda x: x/np.timedelta64(1,'D')) >= 0)) & (df['一码通入住日期'] != '1970-12-30')))
                if (len(df[change_jz])) != 0:
                    df[change_jz].to_excel(
                    rf'./质控名单/12.{sick_type}医学观察方式（改集中）的维护.xlsx',index=False)
                    print(f'{sick_type}医学观察方式（改集中）的维护名单已生成')

            # sick_type = '次密'
            # # 7、关联密接
            # wrong_Associated_contact = ['其他','(空白)','不详','否','0']
            # Associated_contact =  (df['是否排除密接/次密'] != '是') & (df['密接类型'] == '密接的密接') & name_type & (df['地市'] == '广州市') & (df['区县'] == '白云区') & df['关联密接'].str.contains('|'.join(wrong_Associated_contact))
            # df[Associated_contact].to_excel(
            #     rf'./质控名单/7.{sick_type}次密关联密接的维护.xlsx',
            #     index=False)
            # print(f'{sick_type}次密关联密接的维护名单已生成')

    # 清空及创建对应文件夹
    def report_file(self):
        folder_path = './质控名单'
        try:
            shutil.rmtree(folder_path)
        except:
            print('清空失败,可能你本来就清空了或本来就没有这个文件夹~')
        if not os.path.exists(folder_path):
            os.makedirs(folder_path)  # 创建文件夹


if __name__ == '__main__':
    Test = Original()
    print('正在清理上版数据的文件')
    Test.report_file()
    print('清空完毕')

    Test.dealwith()  # 1.数据处理
    for i in range(1):
        name = Test.report['name'].values[i]
        title = Test.report['title'].values[i]
        Test.report_roster()

    print('已完全汇报完毕!~')

